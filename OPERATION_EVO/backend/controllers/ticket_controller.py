import csv
import io
import json
from datetime import datetime
from typing import Dict, List

from database.db import get_db
from services.mistral_service import classify_ticket
from services.similarity_service import build_ai_suggestions
from services.ticket_ai_service import analyze_ticket_problem_solving

ALLOWED_STATUSES = {"ouvert", "en_cours", "resolu"}
ALLOWED_PRIORITIES = {"faible", "normal", "urgent"}
ALLOWED_SEVERITIES = {"faible", "moyenne", "haute", "critique"}
EDITABLE_TICKET_FIELDS = {
    "titre",
    "description",
    "categorie",
    "gravite",
    "priorite",
    "departement_cible",
    "statut",
    "user_id",
    "groupe_id",
}


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _record_activity(conn, ticket_id, action, details=None):
    conn.execute(
        "INSERT INTO ticket_activity (ticket_id, action, details, created_at) VALUES (?, ?, ?, ?)",
        (ticket_id, action, json.dumps(details or {}, ensure_ascii=False), _now()),
    )


def _clean_text(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def _normalize_status(value):
    normalized = _clean_text(value).lower().replace(" ", "_")
    aliases = {"résolu": "resolu", "en-cours": "en_cours", "en__cours": "en_cours"}
    return aliases.get(normalized, normalized)


def validate_ticket_data(data, partial=False):
    errors = {}
    data = data or {}

    if not partial or "titre" in data:
        if not _clean_text(data.get("titre")):
            errors["titre"] = "Le titre est obligatoire."
    if not partial or "description" in data:
        if not _clean_text(data.get("description")):
            errors["description"] = "La description est obligatoire."

    if "statut" in data and data.get("statut") is not None:
        status = _normalize_status(data.get("statut"))
        if status not in ALLOWED_STATUSES:
            errors["statut"] = "Statut invalide."
    if "priorite" in data and data.get("priorite") is not None:
        priority = _clean_text(data.get("priorite")).lower()
        if priority not in ALLOWED_PRIORITIES:
            errors["priorite"] = "Priorité invalide."
    if "gravite" in data and data.get("gravite") is not None:
        severity = _clean_text(data.get("gravite")).lower()
        if severity not in ALLOWED_SEVERITIES:
            errors["gravite"] = "Gravité invalide."

    return errors


def _validate_foreign_keys(conn, user_id=None, group_id=None, user_supplied=False, group_supplied=False):
    if user_supplied and user_id not in (None, ""):
        if conn.execute("SELECT id FROM users WHERE id = ? AND active = 1", (user_id,)).fetchone() is None:
            return {"user_id": "Utilisateur introuvable ou inactif."}
    if group_supplied and group_id not in (None, ""):
        if conn.execute("SELECT id FROM probleme_groupes WHERE id = ?", (group_id,)).fetchone() is None:
            return {"groupe_id": "Groupe de problèmes introuvable."}
    return {}


def analyze_ticket(data):
    """Retourne une analyse IA complète à partir d'un ticket."""
    payload = data or {}
    classification = classify_ticket(payload.get("description", "") or payload.get("titre", ""))
    suggestions = build_ai_suggestions(payload)
    return {
        "message": "analysis ready",
        "classification": classification,
        "ai_suggestions": suggestions,
    }


def analyze_ticket_jira_payload(ticket_id, entity=None):
    result = analyze_ticket_problem_solving(ticket_id=ticket_id, entity=entity)
    if isinstance(result, dict) and result.get("error"):
        return result, 404
    return result


def create_ticket(data, actor_user_id=None):
    """Créer un ticket persistant, avec classification IA et validation minimale."""
    data = data or {}
    errors = validate_ticket_data(data, partial=False)
    if errors:
        return {"error": "validation error", "fields": errors}, 400

    analysis = analyze_ticket(data)
    classification = analysis.get("classification", {})

    status = _normalize_status(data.get("statut", "ouvert"))
    priority = _clean_text(data.get("priorite") or classification.get("priorite") or "normal").lower()
    severity = _clean_text(data.get("gravite") or classification.get("gravite") or "moyenne").lower()
    category = _clean_text(data.get("categorie") or classification.get("categorie") or "autre")
    department = _clean_text(data.get("departement_cible") or classification.get("departement") or "Support")

    normalized_payload = dict(data)
    normalized_payload.update({"statut": status, "priorite": priority, "gravite": severity})
    errors = validate_ticket_data(normalized_payload, partial=False)
    if errors:
        return {"error": "validation error", "fields": errors}, 400

    conn = get_db()
    fk_errors = _validate_foreign_keys(
        conn,
        user_id=data.get("user_id"),
        group_id=data.get("groupe_id"),
        user_supplied="user_id" in data,
        group_supplied="groupe_id" in data,
    )
    if fk_errors:
        conn.close()
        return {"error": "validation error", "fields": fk_errors}, 400

    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO tickets
            (titre, description, categorie, gravite, priorite, departement_cible, statut, user_id, groupe_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            _clean_text(data.get("titre")),
            _clean_text(data.get("description")),
            category,
            severity,
            priority,
            department,
            status,
            data.get("user_id") or None,
            data.get("groupe_id") or None,
        ),
    )
    ticket_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO ticket_history (ticket_id, action, date_action) VALUES (?, ?, ?)",
        (ticket_id, "created", _now()),
    )
    _record_activity(
        conn,
        ticket_id,
        "ticket_created",
        {"source": "dashboard", "actor_user_id": actor_user_id, "classification": classification},
    )
    conn.commit()
    created = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    conn.close()
    return {
        "id": ticket_id,
        "message": "ticket created",
        "ticket": dict(created),
        "classification": classification,
        "ai_suggestions": analysis.get("ai_suggestions"),
    }


def get_tickets(filters=None):
    filters = filters or {}
    conn = get_db()
    query = """
        SELECT t.*, u.nom AS user_nom, u.email AS user_email,
               pg.titre_probleme AS groupe_titre
        FROM tickets t
        LEFT JOIN users u ON u.id = t.user_id
        LEFT JOIN probleme_groupes pg ON pg.id = t.groupe_id
    """
    clauses = []
    params = []
    if filters.get("statut"):
        clauses.append("t.statut = ?")
        params.append(_normalize_status(filters["statut"]))
    if filters.get("priorite"):
        clauses.append("t.priorite = ?")
        params.append(filters["priorite"])
    if filters.get("departement"):
        clauses.append("t.departement_cible = ?")
        params.append(filters["departement"])
    if filters.get("user_id"):
        clauses.append("t.user_id = ?")
        params.append(filters["user_id"])
    if filters.get("groupe_id"):
        clauses.append("t.groupe_id = ?")
        params.append(filters["groupe_id"])
    if clauses:
        query += " WHERE " + " AND ".join(clauses)
    query += " ORDER BY t.id DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_ticket_by_id(ticket_id):
    conn = get_db()
    row = conn.execute(
        """
        SELECT t.*, u.nom AS user_nom, u.email AS user_email,
               pg.titre_probleme AS groupe_titre
        FROM tickets t
        LEFT JOIN users u ON u.id = t.user_id
        LEFT JOIN probleme_groupes pg ON pg.id = t.groupe_id
        WHERE t.id = ?
        """,
        (ticket_id,),
    ).fetchone()
    conn.close()
    return None if row is None else dict(row)


def update_ticket(ticket_id, data, actor_user_id=None):
    data = data or {}
    supplied = {key: data[key] for key in EDITABLE_TICKET_FIELDS if key in data}
    if not supplied:
        return {"error": "no valid fields supplied"}, 400

    if "statut" in supplied and supplied["statut"] is not None:
        supplied["statut"] = _normalize_status(supplied["statut"])
    if "priorite" in supplied and supplied["priorite"] is not None:
        supplied["priorite"] = _clean_text(supplied["priorite"]).lower()
    if "gravite" in supplied and supplied["gravite"] is not None:
        supplied["gravite"] = _clean_text(supplied["gravite"]).lower()
    for field in ("titre", "description", "categorie", "departement_cible"):
        if field in supplied and supplied[field] is not None:
            supplied[field] = _clean_text(supplied[field])
    for field in ("user_id", "groupe_id"):
        if field in supplied and supplied[field] in ("", None):
            supplied[field] = None

    errors = validate_ticket_data(supplied, partial=True)
    if errors:
        return {"error": "validation error", "fields": errors}, 400

    conn = get_db()
    existing = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    if existing is None:
        conn.close()
        return {"error": "ticket not found"}, 404

    fk_errors = _validate_foreign_keys(
        conn,
        user_id=supplied.get("user_id"),
        group_id=supplied.get("groupe_id"),
        user_supplied="user_id" in supplied,
        group_supplied="groupe_id" in supplied,
    )
    if fk_errors:
        conn.close()
        return {"error": "validation error", "fields": fk_errors}, 400

    changed = {}
    for key, value in supplied.items():
        if existing[key] != value:
            changed[key] = {"old": existing[key], "new": value}

    if changed:
        fields = ", ".join(f"{key} = ?" for key in changed)
        values = [changed[key]["new"] for key in changed] + [ticket_id]
        conn.execute(f"UPDATE tickets SET {fields} WHERE id = ?", values)
        conn.execute(
            "INSERT INTO ticket_history (ticket_id, action, date_action) VALUES (?, ?, ?)",
            (ticket_id, "updated", _now()),
        )
        _record_activity(
            conn,
            ticket_id,
            "ticket_updated",
            {"changes": changed, "actor_user_id": actor_user_id},
        )
        if "statut" in changed:
            _record_activity(conn, ticket_id, "status_changed", {"status": changed["statut"]["new"]})
        if "user_id" in changed:
            _record_activity(conn, ticket_id, "assigned", {"user_id": changed["user_id"]["new"]})
        if "priorite" in changed:
            _record_activity(conn, ticket_id, "priority_changed", {"priority": changed["priorite"]["new"]})
        conn.commit()

    updated = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    conn.close()
    return {"message": "ticket updated", "ticket": dict(updated), "changes": changed}


def update_ticket_status(ticket_id, new_status=None, assigned_user_id=None, priority=None, metadata=None):
    """Compatibilité avec les tests et anciens appels de mise à jour partielle."""
    payload = {}
    if new_status is not None:
        payload["statut"] = new_status
    if assigned_user_id is not None:
        payload["user_id"] = assigned_user_id
    if priority is not None:
        payload["priorite"] = priority
    result = update_ticket(ticket_id, payload, actor_user_id=(metadata or {}).get("actor_user_id"))
    if isinstance(result, tuple):
        return result
    ticket = result["ticket"]
    return {
        "id": ticket_id,
        "statut": ticket["statut"],
        "user_id": ticket["user_id"],
        "priorite": ticket["priorite"],
        "message": "status updated",
    }


def assign_ticket_to_user(ticket_id, user_id, group_id=None, actor_user_id=None):
    conn = get_db()
    ticket = conn.execute("SELECT id, groupe_id FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    if ticket is None:
        conn.close()
        return {"error": "ticket not found"}, 404
    if group_id is not None and int(ticket["groupe_id"] or 0) != int(group_id):
        conn.close()
        return {"error": "ticket not found in selected problem group"}, 400
    user = conn.execute("SELECT id FROM users WHERE id = ? AND active = 1", (user_id,)).fetchone()
    if user is None:
        conn.close()
        return {"error": "user not found or inactive"}, 404

    conn.execute("UPDATE tickets SET user_id = ? WHERE id = ?", (user_id, ticket_id))
    conn.execute(
        "INSERT INTO ticket_history (ticket_id, action, date_action) VALUES (?, ?, ?)",
        (ticket_id, "assigned", _now()),
    )
    _record_activity(
        conn,
        ticket_id,
        "assigned",
        {"user_id": user_id, "group_id": group_id, "actor_user_id": actor_user_id},
    )
    conn.commit()
    conn.close()
    return {"id": ticket_id, "user_id": int(user_id), "group_id": group_id, "message": "ticket assigned"}


def get_history_tickets():
    conn = get_db()
    rows = conn.execute(
        """
        SELECT t.*, u.nom AS user_nom, u.email AS user_email
        FROM tickets t
        LEFT JOIN users u ON u.id = t.user_id
        ORDER BY t.id DESC
        """
    ).fetchall()
    conn.close()
    tickets = [dict(row) for row in rows]
    for ticket in tickets:
        status = _normalize_status(ticket.get("statut"))
        ticket["historique_type"] = "traité" if status == "resolu" else "reçu"
    return tickets


def delete_ticket(ticket_id, actor_user_id=None):
    conn = get_db()
    existing = conn.execute("SELECT id FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    if existing is None:
        conn.close()
        return {"error": "ticket not found"}, 404
    conn.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
    conn.commit()
    conn.close()
    return {"id": ticket_id, "message": "ticket deleted", "actor_user_id": actor_user_id}


def create_ticket_comment(ticket_id, user_id, text):
    if not text or not str(text).strip():
        return {"error": "comment text is required"}, 400
    conn = get_db()
    if conn.execute("SELECT id FROM tickets WHERE id = ?", (ticket_id,)).fetchone() is None:
        conn.close()
        return {"error": "ticket not found"}, 404
    if user_id is not None and conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone() is None:
        conn.close()
        return {"error": "user not found"}, 404
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO ticket_comments (ticket_id, user_id, message, date_creation) VALUES (?, ?, ?, ?)",
        (ticket_id, user_id, str(text).strip(), _now()),
    )
    comment_id = cursor.lastrowid
    conn.execute(
        "INSERT INTO ticket_history (ticket_id, action, date_action) VALUES (?, ?, ?)",
        (ticket_id, "comment_added", _now()),
    )
    _record_activity(conn, ticket_id, "comment_added", {"message": str(text).strip(), "user_id": user_id})
    conn.commit()
    conn.close()
    return {"id": comment_id, "ticket_id": ticket_id, "message": str(text).strip(), "user_id": user_id}


def get_ticket_comments(ticket_id):
    conn = get_db()
    rows = conn.execute(
        """
        SELECT c.id, c.ticket_id, c.user_id, c.message, c.date_creation, u.nom AS user_name
        FROM ticket_comments c
        LEFT JOIN users u ON u.id = c.user_id
        WHERE c.ticket_id = ?
        ORDER BY c.id ASC
        """,
        (ticket_id,),
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_ticket_activity(ticket_id):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, ticket_id, action, details, created_at FROM ticket_activity WHERE ticket_id = ? ORDER BY id ASC",
        (ticket_id,),
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_ticket_history(ticket_id):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, ticket_id, action, date_action FROM ticket_history WHERE ticket_id = ? ORDER BY id ASC",
        (ticket_id,),
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def _tickets_export_rows():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, titre, statut, priorite, departement_cible, date_creation FROM tickets ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def _history_export_rows():
    rows = get_history_tickets()
    return [
        {
            "id": row.get("id"),
            "titre": row.get("titre"),
            "statut": row.get("statut"),
            "departement_cible": row.get("departement_cible"),
            "historique_type": row.get("historique_type"),
            "date_creation": row.get("date_creation"),
        }
        for row in rows
    ]


def _csv_from_rows(rows, fieldnames):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({field: row.get(field) for field in fieldnames})
    return output.getvalue()


def _xlsx_from_rows(rows, fieldnames, labels, sheet_name="Export"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        raise RuntimeError("Excel export unavailable: openpyxl not installed") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(labels)
    for row in rows:
        ws.append([row.get(field) for field in fieldnames])

    header_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
    header_font = Font(bold=True, color="7F1D1D")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, field in enumerate(fieldnames, start=1):
        width = max(len(str(labels[idx - 1])), *(len(str(row.get(field) or "")) for row in rows)) if rows else len(str(labels[idx - 1]))
        column_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[column_letter].width = min(max(12, width + 2), 45)

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pdf_from_rows(title, rows, fieldnames, labels):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as exc:
        raise RuntimeError("PDF export unavailable: reportlab not installed") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()

    data = [labels]
    for row in rows:
        data.append([str(row.get(field, "") or "") for field in fieldnames])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )

    story = [Paragraph(title, styles["Heading2"]), Spacer(1, 10), table]
    doc.build(story)
    return buffer.getvalue()


_EXPORT_TABLES = {
    "tickets": {
        "title": "Export des tickets",
        "query": "SELECT id, titre, statut, priorite, departement_cible, date_creation, categorie, gravite, user_id, groupe_id FROM tickets ORDER BY id DESC",
        "fields": ["id", "titre", "statut", "priorite", "departement_cible", "date_creation", "categorie", "gravite", "user_id", "groupe_id"],
        "labels": {
            "id": "ID",
            "titre": "Titre",
            "statut": "Statut",
            "priorite": "Priorite",
            "departement_cible": "Departement",
            "date_creation": "Date",
            "categorie": "Categorie",
            "gravite": "Gravite",
            "user_id": "Utilisateur",
            "groupe_id": "Groupe",
        },
    },
    "history": {
        "title": "Historique des tickets",
        "query": None,
        "fields": ["id", "titre", "statut", "departement_cible", "historique_type", "date_creation"],
        "labels": {
            "id": "ID",
            "titre": "Titre",
            "statut": "Statut",
            "departement_cible": "Departement",
            "historique_type": "Type",
            "date_creation": "Date",
        },
    },
    "users": {
        "title": "Export des utilisateurs",
        "query": "SELECT id, nom, email, departement, role, active FROM users ORDER BY id DESC",
        "fields": ["id", "nom", "email", "departement", "role", "active"],
        "labels": {
            "id": "ID",
            "nom": "Nom",
            "email": "Email",
            "departement": "Departement",
            "role": "Role",
            "active": "Actif",
        },
    },
    "problem_groups": {
        "title": "Export des groupes de problemes",
        "query": "SELECT id, titre_probleme, statut, ticket_maitre_id, date_creation FROM probleme_groupes ORDER BY id DESC",
        "fields": ["id", "titre_probleme", "statut", "ticket_maitre_id", "date_creation"],
        "labels": {
            "id": "ID",
            "titre_probleme": "Titre",
            "statut": "Statut",
            "ticket_maitre_id": "Ticket maitre",
            "date_creation": "Date",
        },
    },
}


def _rows_for_table(table_name: str) -> List[Dict]:
    spec = _EXPORT_TABLES.get(table_name)
    if not spec:
        raise ValueError("unsupported table")

    if table_name == "history":
        return _history_export_rows()

    conn = get_db()
    rows = conn.execute(spec["query"]).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def _normalize_fields(table_name: str, fields):
    available = _EXPORT_TABLES[table_name]["fields"]
    if not fields:
        return available

    if isinstance(fields, str):
        requested = [item.strip() for item in fields.split(",") if item.strip()]
    else:
        requested = [str(item).strip() for item in fields if str(item).strip()]

    normalized = [field for field in requested if field in available]
    return normalized or available


def _apply_filters(rows: List[Dict], filters: Dict[str, str], allowed_fields: List[str]) -> List[Dict]:
    if not filters:
        return rows

    filtered = rows
    for field, value in filters.items():
        if field not in allowed_fields:
            continue
        lookup = str(value or "").strip().lower()
        if not lookup:
            continue
        filtered = [row for row in filtered if lookup in str(row.get(field) or "").lower()]
    return filtered


def _sort_rows(rows: List[Dict], sort_by: str, sort_order: str) -> List[Dict]:
    if not sort_by:
        return rows
    reverse = str(sort_order or "desc").lower() != "asc"

    def key_fn(row):
        value = row.get(sort_by)
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return value
        return str(value).lower()

    return sorted(rows, key=key_fn, reverse=reverse)


def export_custom_table(table="tickets", fields=None, sort_by=None, sort_order="desc", format="excel", filters=None):
    table_name = str(table or "tickets").strip().lower()
    if table_name not in _EXPORT_TABLES:
        raise ValueError("unsupported table")

    selected_fields = _normalize_fields(table_name, fields)
    rows = _rows_for_table(table_name)
    rows = _apply_filters(rows, filters or {}, _EXPORT_TABLES[table_name]["fields"])
    if sort_by in _EXPORT_TABLES[table_name]["fields"]:
        rows = _sort_rows(rows, sort_by=sort_by, sort_order=sort_order)

    labels = [_EXPORT_TABLES[table_name]["labels"].get(field, field) for field in selected_fields]
    fmt = str(format or "excel").lower()

    if fmt == "csv":
        return _csv_from_rows(rows, selected_fields)
    if fmt == "excel":
        return _xlsx_from_rows(rows, selected_fields, labels, sheet_name=table_name)
    if fmt == "pdf":
        return _pdf_from_rows(_EXPORT_TABLES[table_name]["title"], rows, selected_fields, labels)
    raise ValueError("unsupported export format")


def export_tickets(format="excel"):
    fmt = str(format or "excel").lower()
    if fmt == "json":
        return _tickets_export_rows()
    return export_custom_table(
        table="tickets",
        fields=["id", "titre", "statut", "priorite", "departement_cible", "date_creation"],
        sort_by="id",
        sort_order="desc",
        format=fmt,
    )


def export_history_tickets(format="excel"):
    return export_custom_table(
        table="history",
        fields=["id", "titre", "statut", "departement_cible", "historique_type", "date_creation"],
        sort_by="id",
        sort_order="desc",
        format=format,
    )


def export_history_ticket_item(ticket_id, format="excel"):
    fmt = str(format or "excel").lower()
    item = next((row for row in _history_export_rows() if int(row.get("id") or 0) == int(ticket_id)), None)
    if item is None:
        return None

    rows = [item]
    fieldnames = ["id", "titre", "statut", "departement_cible", "historique_type", "date_creation"]
    labels = ["ID", "Titre", "Statut", "Departement", "Type", "Date"]

    if fmt == "csv":
        return _csv_from_rows(rows, fieldnames)
    if fmt == "excel":
        return _xlsx_from_rows(rows, fieldnames, labels, sheet_name=f"ticket_{ticket_id}")
    if fmt == "pdf":
        return _pdf_from_rows(f"Historique ticket #{ticket_id}", rows, fieldnames, labels)
    raise ValueError("unsupported export format")


def get_ticket_metrics():
    conn = get_db()
    rows = conn.execute(
        "SELECT statut, priorite, COUNT(*) as count FROM tickets GROUP BY statut, priorite ORDER BY statut, priorite"
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_ticket_metrics_summary():
    metrics = get_ticket_metrics()
    by_status = {}
    by_priority = {}
    for metric in metrics:
        status = str(metric.get("statut") or "inconnu")
        priority = str(metric.get("priorite") or "inconnu")
        by_status[status] = by_status.get(status, 0) + int(metric.get("count") or 0)
        by_priority[priority] = by_priority.get(priority, 0) + int(metric.get("count") or 0)
    return {
        "total_tickets": sum(by_status.values()),
        "by_status": by_status,
        "by_priority": by_priority,
        "rows": metrics,
    }
